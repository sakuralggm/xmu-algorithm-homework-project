import openpyxl as xlsx
import queue
import random
import math

def load():
    book = xlsx.load_workbook('data.xlsx')
    bidderSheet = book['竞价者信息表']
    # 创建bidderDict，用于储存每个竞价者的信息和出价
    bidders = {}
    for bidder in bidderSheet[2:bidderSheet.max_row]:
        name = ''
        for col in bidder:
            header = bidderSheet.cell(1, col.col_idx).value
            if header == '竞价者':
                bidders[col.value] = {}
                name = col.value
            else:
                bidders[name][header] = col.value
    bidSheet = book['关键词出价表']
    adwords = set([])
    for bid in bidSheet[2:bidSheet.max_row]:
        name = ''
        for col in bid:
            header = bidSheet.cell(1, col.col_idx).value
            if header == '竞价者':
                name = col.value
            else:
                adwords.add(header)
                bidders[name][header] = col.value
    return bidders, list(adwords)

def calculateGrossIncome(bidders, adwords):
    grossIncome = 0
    for adword in adwords:
        pqueue = queue.PriorityQueue()
        for name, info in bidders.items():
            V = info[adword] * info['CTR'] * (1 - math.exp(info['当日已使用金额'] / info['当日总预算'] - 1))
            pqueue.put((-V, (name, info[adword])))
        winner = pqueue.get()
        print(f'关键字 "{adword}" 分配给了{winner[1][0]}，收益{winner[1][1]}元')
        grossIncome += winner[1][1]
        bidders[winner[1][0]]['当日已使用金额'] -= winner[1][1]
    return grossIncome

if __name__ == '__main__':
    data, adwords = load()
    random.shuffle(adwords)
    grossIncome = calculateGrossIncome(data, adwords)
    print(f'最大总收益{grossIncome}元')